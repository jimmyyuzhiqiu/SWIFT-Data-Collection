
import re
import sys
from typing import Optional, Tuple, List, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# =======================
# 配置区：按需修改
# =======================
SWIFT_FILE = r"Z:\To Jimmy Yu\Swift Data Collection\20260108_Swift.xlsx"
SWIFT_SHEET = "Step3_Final"

DW_FILE = r"Z:\To Jimmy Yu\Swift Data Collection\DWCKFS 202512 revised.xlsx"
DW_SHEET = "DWCKFS"

OUTPUT_FILE = r"Z:\To Jimmy Yu\Swift Data Collection\DWCKFS 202512 revised_updated.xlsx"

# 金额模糊匹配阈值：DW金额需落在 [AMT-DELTA, AMT]
AMT_DELTA = 100

# 冲突策略：
# True = 允许覆盖（后来的Step3覆盖DW）
# False = 不覆盖，保留第一次写入，并标橙提示冲突（更安全）
ALLOW_OVERWRITE_ON_CONFLICT = False


# =======================
# 工具函数
# =======================
def excel_col_letter_to_index(letter: str) -> int:
    """Excel列字母 -> 0-based index, 如 'A'->0, 'X'->23"""
    letter = letter.upper().strip()
    idx = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"非法列字母: {letter}")
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


def normalize_account(x) -> str:
    """账号统一成字符串，去空格；尽力处理科学计数法"""
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if re.fullmatch(r"-?\d+(\.\d+)?[eE]\+?\d+", s):
        try:
            s = format(int(float(s)), "d")
        except Exception:
            pass
    return s


def to_number(x) -> Optional[float]:
    """把金额转成 float（支持带逗号/空格），失败返回 None"""
    if pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def pick_column(df: pd.DataFrame, preferred_name: str, fallback_excel_letter: str) -> pd.Series:
    """优先按列名取，否则按Excel列字母取（0-based）"""
    if preferred_name in df.columns:
        return df[preferred_name]
    idx = excel_col_letter_to_index(fallback_excel_letter)
    if idx < 0 or idx >= df.shape[1]:
        raise KeyError(f"找不到列名'{preferred_name}'且fallback列 {fallback_excel_letter} 超出范围。")
    return df.iloc[:, idx]


def find_best_by_amount(amount_list: List[Tuple[float, int]], target_amt: float, delta: float) -> Optional[int]:
    """
    在amount_list中找落在 [target-delta, target] 的记录，返回最接近target的dw_df行号(index)
    """
    low = target_amt - delta
    high = target_amt
    candidates = [(abs(target_amt - a), idx) for a, idx in amount_list if low <= a <= high]
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])  # 距离最小优先
    return candidates[0][1]


def build_dw_indexes(dw_df: pd.DataFrame) -> Tuple[Dict[str, List[int]], List[Tuple[float, int]]]:
    """
    构建两个索引：
    1) account -> [dw_df_index...]
    2) amount_list: [(amount_value, dw_df_index), ...]
    """
    dw_acc_col = pick_column(dw_df, "交易对手存款账户编码", "X")
    dw_amt_col = pick_column(dw_df, "存款发生金额", "O")

    account_map: Dict[str, List[int]] = {}
    for i, v in dw_acc_col.items():
        acc = normalize_account(v)
        if acc:
            account_map.setdefault(acc, []).append(i)

    amount_list = []
    for i, v in dw_amt_col.items():
        amt = to_number(v)
        if amt is not None:
            amount_list.append((amt, i))

    return account_map, amount_list


def locate_dw_target_col(ws_dw) -> int:
    """
    定位DW需要写入的列：
    优先表头中找 '交易对手账户开户行号'
    否则用 Y 列（第25列）
    """
    header_row = 1
    header_map = {}
    for c in range(1, ws_dw.max_column + 1):
        v = ws_dw.cell(header_row, c).value
        if v is None:
            continue
        header_map[str(v).strip()] = c

    if "交易对手账户开户行号" in header_map:
        return header_map["交易对手账户开户行号"]
    else:
        return excel_col_letter_to_index("Y") + 1  # openpyxl是1-based


# =======================
# 主流程
# =======================
def main():
    print("开始处理（修正版：生成新的DW文件，并把DW的Y列/交易对手账户开户行号改为Step3的CP SWIFT）...")

    # 1) 读取Step3_Final
    step_df = pd.read_excel(SWIFT_FILE, sheet_name=SWIFT_SHEET, engine="openpyxl").reset_index(drop=True)
    for need in ["CP A/C", "AMT", "CP SWIFT"]:
        if need not in step_df.columns:
            raise KeyError(f"Step3_Final缺少列: {need}。实际列：{list(step_df.columns)}")

    # 2) 读取DW（pandas用于匹配）
    dw_df = pd.read_excel(DW_FILE, sheet_name=DW_SHEET, engine="openpyxl").reset_index(drop=True)
    account_map, amount_list = build_dw_indexes(dw_df)

    # 3) openpyxl加载DW原工作簿（用于写回并尽量保留格式）
    wb_dw = load_workbook(DW_FILE)
    if DW_SHEET not in wb_dw.sheetnames:
        raise KeyError(f"DW文件中找不到sheet: {DW_SHEET}")
    ws_dw = wb_dw[DW_SHEET]

    target_col = locate_dw_target_col(ws_dw)

    # 标色
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # 冲突
    bold_font = Font(bold=True)

    # 4) 开始匹配并写回DW
    matched_ac = 0
    matched_amt = 0
    written = 0
    conflicts = []

    # 记录：dw_row_index -> 已写入的swift值
    dw_written_value: Dict[int, str] = {}

    unmatched_rows = []

    for i in range(len(step_df)):
        cp_ac = normalize_account(step_df.at[i, "CP A/C"])
        amt = to_number(step_df.at[i, "AMT"])
        cp_swift = "" if pd.isna(step_df.at[i, "CP SWIFT"]) else str(step_df.at[i, "CP SWIFT"]).strip()

        # Step3的Excel行号（假设第1行表头）
        step_excel_row = i + 2

        if not cp_swift:
            unmatched_rows.append((step_excel_row, cp_ac, amt, cp_swift, "CP SWIFT为空，未写回"))
            continue

        dw_hit = None
        hit_type = None

        # 账号优先
        if cp_ac and cp_ac in account_map:
            dw_hit = account_map[cp_ac][0]
            hit_type = "AC"

        # 金额模糊匹配
        if dw_hit is None and amt is not None:
            dw_hit = find_best_by_amount(amount_list, amt, AMT_DELTA)
            if dw_hit is not None:
                hit_type = "AMT"

        if dw_hit is None:
            unmatched_rows.append((step_excel_row, cp_ac, amt, cp_swift, "未匹配到DW"))
            continue

        # 写回DW：dw_hit 是 dw_df 的行号（0-based），对应Excel行号=dw_hit+2
        dw_excel_row = dw_hit + 2
        cell = ws_dw.cell(row=dw_excel_row, column=target_col)

        # 冲突检查
        if dw_hit in dw_written_value and dw_written_value[dw_hit] != cp_swift:
            conflicts.append((dw_excel_row, dw_written_value[dw_hit], cp_swift))
            cell.fill = orange_fill
            if ALLOW_OVERWRITE_ON_CONFLICT:
                cell.value = cp_swift
                dw_written_value[dw_hit] = cp_swift
            # 若不允许覆盖，就保持第一次写入不动
        else:
            # 第一次写入 or 同值重复
            cell.value = cp_swift
            dw_written_value[dw_hit] = cp_swift
            written += 1
            if hit_type == "AC":
                matched_ac += 1
            else:
                matched_amt += 1

    # 5) 把未匹配行写入新sheet，并标黄
    sheet_name_unmatched = "Unmatched_Step3"
    if sheet_name_unmatched in wb_dw.sheetnames:
        del wb_dw[sheet_name_unmatched]
    ws_un = wb_dw.create_sheet(sheet_name_unmatched)

    headers = ["Step3_Excel行号", "CP A/C", "AMT", "CP SWIFT", "原因"]
    for c, h in enumerate(headers, 1):
        ws_un.cell(row=1, column=c, value=h).font = bold_font

    for r, row in enumerate(unmatched_rows, start=2):
        for c, v in enumerate(row, start=1):
            ws_un.cell(row=r, column=c, value=v)
            ws_un.cell(row=r, column=c).fill = yellow_fill

    # 6) 保存新DW文件
    wb_dw.save(OUTPUT_FILE)

    # 7) 终端汇总输出
    print("\n==================== 处理完成（生成DW新文件） ====================")
    print(f"输出文件: {OUTPUT_FILE}")
    print(f"账号匹配写回: {matched_ac}")
    print(f"金额模糊匹配写回: {matched_amt}")
    print(f"实际写入DW单元格次数(去重后): {written}")
    print(f"匹配失败/未写回(见Unmatched_Step3并已标黄): {len(unmatched_rows)}")
    print(f"冲突数(同一DW行匹配到不同CP SWIFT，DW目标单元格标橙): {len(conflicts)}")

    if unmatched_rows:
        print("\n--- 未匹配/未写回（前50条）---")
        for k, (rownum, ac, amt, swift, reason) in enumerate(unmatched_rows[:50], 1):
            print(f"{k:02d}. Step3行={rownum} | CP A/C={ac} | AMT={amt} | CP SWIFT={swift} | 原因={reason}")
        if len(unmatched_rows) > 50:
            print(f"... 还有 {len(unmatched_rows) - 50} 条未显示（详见输出文件Unmatched_Step3）")

    if conflicts:
        print("\n--- 冲突（前50条）---")
        for k, (dw_row, old_v, new_v) in enumerate(conflicts[:50], 1):
            print(f"{k:02d}. DW行={dw_row} | 已写入={old_v} | 新匹配={new_v} | 处理策略={'覆盖' if ALLOW_OVERWRITE_ON_CONFLICT else '保留第一次'}")
        if len(conflicts) > 50:
            print(f"... 还有 {len(conflicts) - 50} 条未显示")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n程序出错: {e}")
        sys.exit(1)
